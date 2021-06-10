from myczsc.myanalyze import My_CZSC
import os
from data.tdx import get_data_from_tdxfile
from datetime import datetime
from typing import List
from czsc.data.jq import get_kline, get_index_stocks
import czsc
from czsc.analyze import RawBar
from czsc.enum import Signals
import traceback
import baostock as bs
import pandas as pd
import tushare as ts
import struct
import openpyxl
from openpyxl.reader.excel import load_workbook

TDX_DIR = r"D:\new_tdxqh"  # 首先要设置通达信的安装目录


def is_buy(bars: List[RawBar]) -> {}:
    """判断一个股票现在是否有买入信号，没有则返回{}
    :param 
    :return:形如{'B1':'X9B1','ubil':3}
    """
    c = My_CZSC(bars, freq="日线")
    signals = {}
    # 在这里判断是否有买入形态
    if c.signals['B1'] in ['X9B1', 'X11B1', 'X13B1']:
        signals['B1'] = c.signals['B1']
    if c.signals['B2'] in ['X9B1', 'X11B1', 'X13B1']:
        signals['B2'] = c.signals['B2']
    if c.signals['倒1形态'] in [Signals.LI0.value]:
        signals['B3'] = 'B3'
    if len(signals) != 0:
        signals['ubil'] = c.signals['未完成笔长度']
    return signals


def write_excel(signals_list):
    '''
    :param signals_list形如[{'000001.sz':{'B1':'B1','ubil':3}},{'000002.sz':{'B3':'B3','ubil':5}}]
    :return:
    '''
    addr = 'sele.xlsx'
    cur_date = datetime.now().strftime('%Y-%m-%d')
    #文件是否存在，若不存在创建，并初始化sheet，若存在打开文件写入信号
    if os.path.exists(addr) and os.path.isfile(addr):
        wb = openpyxl.load_workbook(addr)
    else:
        wb = openpyxl.Workbook(addr)
        wb.create_sheet('B1')
        wb.create_sheet('B2')
        wb.create_sheet('B3')
        wb.create_sheet('B23')
    for i in signals_list:  #signals_list是列表类型，每只证券的信号为signals，是map类型
        symbol = list(i.keys())[0]
        signals = i[symbol]  #signals是map类型
        row = [cur_date, symbol]
        if len(signals) > 2:  #多个信号出现
            sheet = wb['B23']
            row.append(
                signals.get('B1', '') + signals.get('B2', '') +
                signals.get('B3', ''))
        elif signals.get('B1', '').endswith('B1'):  #信号是一买
            sheet = wb['B1']
            row.append(signals.get('B1'))
        elif signals.get('B2', '')[-2:] == 'B1':  #信号是二买
            sheet = wb['B2']
            row.append(signals.get('B2'))
        elif signals.get('B3') == 'B3':  #信号是三买
            sheet = wb['B3']
            row.append('B3')
        row.append(signals['ubil'])
        sheet.append(row)
    wb.save(addr)
    print("写入数据成功！")


def select():
    # 找出沪市6开头的，中三买的票
    rootdir = TDX_DIR + r"\vipdoc\sh\lday"
    signals_list = []  #用来存储所有证券的信号，一只证券可能存在多个信号
    list = os.listdir(rootdir)  # 列出文件夹下所有的目录与文件
    for i in range(0, len(list)):
        try:
            symbol = list[i][2:8]
            # if symbol.startswith("60000"):
            bars = get_data_from_tdxfile(symbol, 'sh')
            signals = is_buy(bars)
            if len(signals) > 0:  #若存在信号，把证券及其信号，存入signals_list
                signals_list.append({symbol + '.sh': signals})
                print("{}{} - {}".format('sh', symbol, signals))
        except Exception as e:
            traceback.print_exc()
            print("{} 执行失败 - {}".format(symbol, e))

    # 找出深圳中0开头的三买的票
    rootdir = TDX_DIR + r"\vipdoc\sz\lday"
    list = os.listdir(rootdir)  # 列出文件夹下所有的目录与文件
    for i in range(0, len(list)):
        try:
            symbol = list[i][2:8]
            # if symbol.startswith("00000"):
            bars = get_data_from_tdxfile(symbol, 'sz')
            signals = is_buy(bars)
            if len(signals) > 0:
                signals_list.append({symbol + '.sz': signals})
                print("{}{} - {}".format('sz', symbol, signals))
        except Exception as e:
            traceback.print_exc()
            print("{} 执行失败 - {}".format(symbol, e))

    write_excel(signals_list)  #把信号列表，写入excel


if __name__ == '__main__':
    select()