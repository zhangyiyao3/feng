from collections import OrderedDict
from factors.stock import get_finance_factors
from myczsc.myanalyze1 import My_CZSC
import os
from data.tdx import get_data_from_tdxfile
from datetime import datetime
from typing import List
from czsc.data.jq import get_kline, get_index_stocks
import czsc
from czsc.analyze import RawBar
from czsc.enum import Signals
import traceback
import baostock as bs
import pandas as pd
import tushare as ts
import struct
import openpyxl
from openpyxl.reader.excel import load_workbook

TDX_DIR = r"D:\new_tdxqh"  # 首先要设置通达信的安装目录


def is_buy(market, symbol, bars: List[RawBar]) -> {}:
    """判断一个股票现在是否有买入信号，没有则返回{}
    :param 
    :return:真/假,因子字典
    因子字典形如{'symbol':'sh600001','date':'2021-06-08',:,'B1':'X9B1','ubil':3}
    """
    factors = OrderedDict({})

    c = My_CZSC(bars, freq="日线")
    buy_flg = False
    # 在这里判断是否有买入形态
    if c.signals['B1'] in ['X9B1', 'X11B1', 'X13B1']:
        factors['date'] = bars[-1].dt.strftime('%Y-%m-%d %H:%M')
        factors['symbol'] = market + symbol
        # factors['name'] = market + symbol
        factors['B1'] = c.signals['B1']
        buy_flg = True
    if c.signals['B2'] in ['X9B1', 'X11B1', 'X13B1']:
        factors['date'] = bars[-1].dt.strftime('%Y-%m-%d %H:%M')
        factors['symbol'] = market + symbol
        # factors['name'] = market + symbol
        factors['B2'] = c.signals['B2']
        buy_flg = True
    if c.signals['倒1形态'] in [Signals.LI0.value]:
        factors['date'] = bars[-1].dt.strftime('%Y-%m-%d %H:%M')
        factors['symbol'] = market + symbol
        # factors['name'] = market + symbol
        factors['B3'] = 'B3'
        buy_flg = True
    if buy_flg:
        factors['ubil'] = c.signals['未完成笔长度']
        factors.update(get_finance_factors(market, symbol, bars[-1].close))
        # factors.update(get_tec_factors())
    return buy_flg, factors


def write_excel(df):
    '''
    :param signals_list形如[{'000001.sz':{'B1':'B1','ubil':3}},{'000002.sz':{'B3':'B3','ubil':5}}]
    :return:
    '''
    cur_date = datetime.now().strftime('%Y-%m-%d')
    addr = '.\export\sele' + cur_date + '.xlsx'
    #文件是否存在，若不存在创建，并初始化sheet，若存在打开文件写入信号
    if os.path.exists(addr) and os.path.isfile(addr):
        wb = openpyxl.load_workbook(addr)
    else:
        wb = openpyxl.Workbook(addr)
        title = [
            'date', 'symbol', 'name', '买卖点', '信号出现Days', '二级行业', '地域', '总市值/亿',
            'ROE%', 'ROA%', '毛利率%', '净利率%', '毛利增长率%', '收入增长率%', '资产负债率%',
            '营业收入/亿', 'PEG', '动PE', '静PE', 'PB', 'PS', '现金流/净利润', '经营cash/亿',
            '总cash/亿', 'ma由高至低(含股价)', '向上ma', '向下ma', '信号至今涨跌%', '信号至今日均涨跌%',
            '20抵扣价5D方向', '60抵扣价5D方向', '120抵扣价5D方向', '250抵扣价5D方向'
        ]
        wb.create_sheet('B1')
        sheet = wb['B1']
        sheet.append(title)
        wb.create_sheet('B2')
        sheet = wb['B2']
        sheet.append(title)
        wb.create_sheet('B3')
        sheet = wb['B3']
        sheet.append(title)
        wb.create_sheet('B23')
        sheet = wb['B23']
        sheet.append(title)
    for i in df:  #signals_list是列表类型，每只证券的信号为signals，是map类型
        symbol = list(i.keys())[0]
        signals = i[symbol]  #signals是map类型
        row = [cur_date, symbol]
        if len(signals) > 2:  #多个信号出现
            sheet = wb['B23']
            row.append(
                signals.get('B1', '') + signals.get('B2', '') +
                signals.get('B3', ''))
        elif signals.get('B1', '').endswith('B1'):  #信号是一买
            sheet = wb['B1']
            row.append(signals.get('B1'))
        elif signals.get('B2', '')[-2:] == 'B1':  #信号是二买
            sheet = wb['B2']
            row.append(signals.get('B2'))
        elif signals.get('B3') == 'B3':  #信号是三买
            sheet = wb['B3']
            row.append('B3')
        row.append(signals['ubil'])
        row.append(signals['行业'])
        row.append(signals['地域'])
        row.append(signals['总市值-亿'])
        row.append(signals['经营现金流-亿'])
        row.append(signals['总现金流-亿'])
        row.append(signals['ROE%'])
        row.append(signals['净利率%'])
        row.append(signals['毛利率%'])
        row.append(signals['资产负债率%'])
        row.append(signals['PE'])
        row.append(signals['PB'])
        sheet.append(row)
    wb.save(addr)
    print("写入数据成功！")

def df_excel(df):
    '''
    :param df形如[{'000001.sz':{'B1':'B1','ubil':3}},{'000002.sz':{'B3':'B3','ubil':5}}]
    :return:
    '''
    cur_date = datetime.now().strftime('%Y-%m-%d')
    addr = '.\export\sele' + cur_date + '.xlsx'
    #文件是否存在，若不存在创建，并初始化sheet，若存在打开文件写入信号
    df.to_excel(addr, encoding='utf-8', index=False, header=True)
    print("写入数据成功！")

def select():
    # 找出沪市6开头的，中三买的票
    rootdir = TDX_DIR + r"\vipdoc\sh\lday"
    df_factors = pd.DataFrame()  #用来存储所有证券的信号，一只证券可能存在多个信号
    pd.set_option('precision', 2)
    # df_factors.round(2)
    list = os.listdir(rootdir)  # 列出文件夹下所有的目录与文件
    for i in range(0, len(list)):
        try:
            symbol = list[i][2:8]
            # if symbol.startswith("60001"):
            bars = get_data_from_tdxfile(symbol, 'sh')
            buy_flg, factors = is_buy('sh', symbol, bars)
            if buy_flg:  #若存在信号，把证券及其信号，存入signals_list
                df_factors = df_factors.append(factors, ignore_index=True)
                # print("{}{} - {}".format('sh', symbol, factors))
        except Exception as e:
            traceback.print_exc()
            print("{} 执行失败 - {}".format(symbol, e))

    # 找出深圳中0开头的三买的票
    rootdir = TDX_DIR + r"\vipdoc\sz\lday"
    list = os.listdir(rootdir)  # 列出文件夹下所有的目录与文件
    for i in range(0, len(list)):
        try:
            symbol = list[i][2:8]
            # if symbol.startswith("300870") or symbol.startswith("163415"):
            bars = get_data_from_tdxfile(symbol, 'sz')
            buy_flg, factors = is_buy('sz', symbol, bars)
            if buy_flg:
                df_factors = df_factors.append(factors, ignore_index=True)
                # print("{}{} - {}".format('sz', symbol, factors))
        except Exception as e:
            traceback.print_exc()
            print("{} 执行失败 - {}".format(symbol, e))
    # print(df_factors)
    df_excel(df_factors)  #把信号列表，写入excel


if __name__ == '__main__':
    # print(get_finance_factors('sh', '600816', 5))
    # print(get_finance_factors('sh', '515220', 5))
    select()
