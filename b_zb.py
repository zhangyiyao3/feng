from datetime import datetime
from typing import List
from czsc.data.jq import get_kline, get_index_stocks
import czsc
from czsc.analyze import CZSC, RawBar
from czsc.enum import Signals
# 如果需要实盘行情，导入聚宽数据
from czsc.data.jq import *
import traceback
import baostock as bs
import pandas as pd
import tushare as ts
import struct
import openpyxl
from openpyxl.reader.excel import load_workbook

assert czsc.__version__ == '0.7.2'
TDX_DIR = r"D:\new_tdxqh"  # 首先要设置通达信的安装目录


# 使用第三方数据，只需要定义一个K线转换函数
def format_kline_ts(kline: pd.DataFrame) -> List[RawBar]:
    """
    :param kline: Tushare 数据接口返回的K线数据
    :return: 转换好的K线数据
    """
    bars = []
    records = kline.to_dict('records')
    for record in records:
        # 将每一根K线转换成 RawBar 对象
        bar = RawBar(symbol=record['ts_code'],
                     dt=pd.to_datetime(record['trade_date']),
                     open=record['open'],
                     close=record['close'],
                     high=record['high'],
                     low=record['low'],
                     vol=record['vol'])
        bars.append(bar)
    return bars


# 使用第三方数据，只需要定义一个K线转换函数
def format_kline_bs(kline: pd.DataFrame) -> List[RawBar]:
    """
    :param kline: Tushare 数据接口返回的K线数据
    :return: 转换好的K线数据
    """
    bars = []
    records = kline.to_dict('records')
    for record in records:
        # 将每一根K线转换成 RawBar 对象
        bar = RawBar(symbol=record['ts_code'],
                     dt=pd.to_datetime(record['trade_date']),
                     open=record['open'],
                     close=record['close'],
                     high=record['high'],
                     low=record['low'],
                     vol=record['vol'])
        bars.append(bar)
    return bars


def is_buy(bars: List[RawBar]) -> List:
    """判断一个股票现在是否有买入信号，没有则返回[]"""
    c = CZSC(bars, freq="日线")
    signals = []
    # 在这里判断是否有买入形态
    if c.signals['倒1形态'] in [Signals.LA0.value]:
        signals.append('B1')
    if c.signals['倒3形态'] in [Signals.LA0.value]:
        signals.append('B2')
    if c.signals['倒1形态'] in [Signals.LI0.value]:
        signals.append('B3')
    return signals


def tshare():
    # 获取上证50最新成分股列表，这里可以换成自己的股票池
    symbols: List = ["000985.SH"]
    # symbols: List = get_index_stocks("000985.XSHG")
    for symbol in symbols:
        try:
            # 调用tushare的K线获取方法，Tushare数据的使用方法，请参考：https://tushare.pro/document/2
            end_date = datetime.now()
            start_date = end_date - timedelta(days=1000)
            df = ts.pro_bar(ts_code=symbol,
                            adj='qfq',
                            asset="E",
                            start_date=start_date.strftime("%Y%m%d"),
                            end_date=end_date.strftime("%Y%m%d"))
            bars = format_kline_ts(df)
            signals = is_buy(bars)
            if len(signals) > 0:
                print("{} - {}".format(symbol, signals))
        except Exception as e:
            traceback.print_exc()
            print("{} 执行失败 - {}".format(symbol, e))


def jquan():
    # 获取上证50最新成分股列表，这里可以换成自己的股票池
    symbols: List = ["000985.XSHG"]
    # symbols: List = get_index_stocks("000985.XSHG")
    for symbol in symbols:
        try:
            bars = get_kline(symbol,
                             freq="D",
                             end_date=datetime.now(),
                             count=1000)
            signals = is_buy(bars)
            if len(signals) > 0:
                print("{} - {}".format(symbol, signals))
        except Exception as e:
            traceback.print_exc()
            print("{} 执行失败 - {}".format(symbol, e))
    print("聚宽剩余调用次数：{}".format(get_query_count()))


def bstock():
    bs.login()
    # 获取指定日期的指数、股票列表信息
    end_date = datetime.now()
    e_date = end_date - timedelta(days=10)
    start_date = end_date - timedelta(days=1000)
    start_date, end_date = datetime.strftime(
        start_date, '%Y-%m-%d'), end_date.strftime('%Y-%m-%d')
    stock_rs = bs.query_all_stock(e_date.strftime('%Y-%m-%d'))
    stock_df = stock_rs.get_data()
    # print(stock_df["code"])
    #针对股票列表，逐个处理：获取历史数据生成dataframe、格式化、传入
    # for symbol in ["sh.000001",'sh.000002']:
    bars = []
    for symbol in stock_df["code"]:
        try:
            pd.m
            data_df = pd.DataFrame()
            k_rs = bs.query_history_k_data_plus(
                symbol, "date,code,open,high,low,close,volume", start_date,
                end_date)
            data_df = data_df.append(k_rs.get_data())
            #格式化bars
            records = data_df.to_dict('records')
            for record in records:
                # 将每一根K线转换成 RawBar 对象
                bar = RawBar(symbol=record['code'],
                             dt=pd.to_datetime(record['date']),
                             open=float(record['open']),
                             close=float(record['close']),
                             high=float(record['high']),
                             low=float(record['low']),
                             vol=float(record['volume']))
                bars.append(bar)
            signals = is_buy(bars)
            if len(signals) > 0:
                print("{} - {}".format(symbol, signals))
        except Exception as e:
            traceback.print_exc()
            print("{} 执行失败 - {}".format(symbol, e))
        finally:
            bars.clear()
    bs.logout()


def bb():
    # 登陆系统
    lg = bs.login()
    # 显示登陆返回信息
    print('login respond error_code:' + lg.error_code)
    print('login respond  error_msg:' + lg.error_msg)

    # 获取中证500成分股
    rs = bs.query_zz500_stocks()
    print('query_zz500 error_code:' + rs.error_code)
    print('query_zz500  error_msg:' + rs.error_msg)

    # 打印结果集
    zz500_stocks = []
    while (rs.error_code == '0') & rs.next():
        # 获取一条记录，将记录合并在一起
        zz500_stocks.append(rs.get_row_data())
    result = pd.DataFrame(zz500_stocks, columns=rs.fields)
    # 结果集输出到csv文件
    result.to_csv("D:/zz500_stocks.csv", encoding="gbk", index=False)
    print(result)

    # 登出系统
    bs.logout()


def download_data():
    bs.login()
    # 获取指定日期的指数、股票数据
    stock_rs = bs.query_all_stock()
    stock_df = stock_rs.get_data()
    data_df = pd.DataFrame()
    end_date = datetime.now()
    start_date = end_date - timedelta(days=1000)
    start_date, end_date = datetime.strftime(
        start_date, '%Y-%m-%d'), end_date.strftime('%Y-%m-%d')
    for code in ["sh.600000"]:
        # for code in stock_df["code"]:
        k_rs = bs.query_history_k_data_plus(
            code, "date,code,open,high,low,close,volume", start_date, end_date)
        data_df = data_df.append(k_rs.get_data())
    bars = []
    records = data_df.to_dict('records')
    for record in records:
        # 将每一根K线转换成 RawBar 对象
        bar = RawBar(symbol=record['code'],
                     dt=pd.to_datetime(record['date']),
                     open=record['open'],
                     close=record['close'],
                     high=record['high'],
                     low=record['low'],
                     vol=record['volume'])
        bars.append(bar)
    bs.logout()
    print(bars)


# 从通达信目录读入数据
def get_data_from_tdxfile(stock_code, type) -> List[RawBar]:
    '''
    stock_code:股票代码 600667
    type：市场代码，sh沪市，sz深市
    '''
    bars = []
    filepath = TDX_DIR + r'\vipdoc\\' + type + r'\lday\\' + type + stock_code + '.day'
    with open(filepath, 'rb') as f:
        while True:
            stock_date = f.read(4)
            stock_open = f.read(4)
            stock_high = f.read(4)
            stock_low = f.read(4)
            stock_close = f.read(4)
            stock_amount = f.read(4)
            stock_vol = f.read(4)
            stock_reservation = f.read(4)
            if not stock_date:
                break
            stock_date = struct.unpack("l", stock_date)  # 4字节 如20091229
            stock_open = struct.unpack("l", stock_open)  # 开盘价*100
            stock_high = struct.unpack("l", stock_high)  # 最高价*100
            stock_low = struct.unpack("l", stock_low)  # 最低价*100
            stock_close = struct.unpack("l", stock_close)  # 收盘价*100
            stock_amount = struct.unpack("f", stock_amount)  # 成交额
            stock_vol = struct.unpack("l", stock_vol)  # 成交量
            stock_reservation = struct.unpack("l", stock_reservation)  # 保留值
            date_format = datetime.strptime(str(stock_date[0]),
                                            '%Y%M%d')  # 格式化日期
            date_format = date_format.strftime('%Y-%M-%d')

            bar = RawBar(symbol=stock_code,
                         dt=pd.to_datetime(date_format),
                         open=stock_open[0] / 100,
                         close=stock_close[0] / 100.0,
                         high=stock_high[0] / 100.0,
                         low=stock_low[0] / 100.0,
                         vol=stock_vol[0])
            bars.append(bar)
        return bars


def tdx():
    # 找出沪市6开头的，中三买的票
    rootdir = TDX_DIR + r"\vipdoc\sh\lday"

    list = os.listdir(rootdir)  # 列出文件夹下所有的目录与文件
    # print(list)
    for i in range(0, len(list)):
        try:
            scode = list[i][2:8]
            # if scode.startswith("6"):
            bars = get_data_from_tdxfile(scode, 'sh')
            signals = is_buy(bars)
            if len(signals) > 0:
                print("{}{} - {}".format('sh', scode, signals))
        except Exception as e:
            traceback.print_exc()
            print("{} 执行失败 - {}".format(scode, e))

    # 找出深圳中0开头的三买的票
    rootdir = TDX_DIR + r"\vipdoc\sz\lday"
    list = os.listdir(rootdir)  # 列出文件夹下所有的目录与文件
    for i in range(0, len(list)):
        try:
            scode = list[i][2:8]
            # if scode.startswith("0"):
            bars = get_data_from_tdxfile(scode, 'sz')
            signals = is_buy(bars)
            if len(signals) > 0:
                print("{}{} - {}".format('sz', scode, signals))
        except Exception as e:
            traceback.print_exc()
            print("{} 执行失败 - {}".format(scode, e))


def tdx_excel():
    # 找出沪市6开头的，中三买的票
    rootdir = TDX_DIR + r"\vipdoc\sh\lday"
    signals_list = []  #用来存储所有证券的信号，一只证券可能存在多个信号
    list = os.listdir(rootdir)  # 列出文件夹下所有的目录与文件
    for i in range(0, len(list)):
        try:
            symbol = list[i][2:8]
            # if symbol.startswith("60000"):
            bars = get_data_from_tdxfile(symbol, 'sh')
            signals = is_buy(bars)
            if len(signals) > 0:  #若存在信号，把证券及其信号，存入signals_list
                signals_list.append({symbol + '.sh': signals})
                print("{}{} - {}".format('sh', symbol, signals))
        except Exception as e:
            traceback.print_exc()
            print("{} 执行失败 - {}".format(symbol, e))

    # 找出深圳中0开头的三买的票
    rootdir = TDX_DIR + r"\vipdoc\sz\lday"
    list = os.listdir(rootdir)  # 列出文件夹下所有的目录与文件
    for i in range(0, len(list)):
        try:
            symbol = list[i][2:8]
            # if symbol.startswith("00000"):
            bars = get_data_from_tdxfile(symbol, 'sz')
            signals = is_buy(bars)
            if len(signals) > 0:
                signals_list.append({symbol + '.sz': signals})
                print("{}{} - {}".format('sz', symbol, signals))
        except Exception as e:
            traceback.print_exc()
            print("{} 执行失败 - {}".format(symbol, e))
    write_excel(signals_list)  #把信号列表，写入excel


def write_excel(signals_list):
    '''
    :param sheet:sheet的名称
    :return:
    '''
    addr = 'zb.xlsx'
    #文件是否存在，若不存在创建，并初始化sheet，若存在打开文件写入信号
    if os.path.exists(addr) and os.path.isfile(addr):
        wb = openpyxl.load_workbook(addr)
    else:
        wb = openpyxl.Workbook(addr)
        wb.create_sheet('B1')
        wb.create_sheet('B2')
        wb.create_sheet('B3')
        wb.create_sheet('B23')
    for i in signals_list:  #每只证券的信号为signals，是map类型
        cur_date = datetime.now().strftime('%Y-%m-%d')
        symbol = list(i.keys())[0]
        signals = i[symbol]  #signals是list类型
        if len(signals) > 1:
            sheet = wb['B23']
            row = [cur_date, symbol]
            row.extend(signals)
            sheet.append(row)
        elif signals[0] == 'B1':  #信号是一买
            sheet = wb['B1']
            sheet.append([cur_date, symbol, 'B1'])
        elif signals[0] == 'B2':  #信号是二买
            sheet = wb['B2']
            sheet.append([cur_date, symbol, 'B2'])
        elif signals[0] == 'B3':  #信号是三买
            sheet = wb['B3']
            sheet.append([cur_date, symbol, 'B3'])
    wb.save(addr)
    print("写入数据成功！")


if __name__ == '__main__':
    # print('a')
    # tdx()
    tdx_excel()
    # tshare()
    # jquan()
    # bstock()
    # bb()
    # download_data()
